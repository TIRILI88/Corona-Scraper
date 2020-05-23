import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime



def NTV():
  url = 'https://www.focus.de/'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=False)
  soup = BeautifulSoup(r.content, 'lxml')
  words = soup.findAll(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []
  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)
  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url, count, the_word))
  return count

def BBC():
  url = 'https://www.bbc.com/'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=False)
  soup = BeautifulSoup(r.content, 'lxml')
  words = soup.find_all(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []
  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)
  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url,count, the_word))
  return count

def CNN():
  url = 'https://www.cnn.com/'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=False)
  soup = BeautifulSoup(r.text, 'html.parser')
  words = soup.find_all(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []
  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)
  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url, count, the_word))
  return count

def googleNewsUSA():
  url = 'https://news.google.com/topstories?hl=en-US&gl=US&ceid=US:en'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=True)
  soup = BeautifulSoup(r.text, 'html.parser')
  words = soup.find_all(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []

  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)
  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url, count, the_word))
  return count

def googleNewsDE():
  url = 'https://news.google.com/?hl=de&gl=DE&ceid=DE:de'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=True)
  soup = BeautifulSoup(r.text, 'html.parser')
  words = soup.find_all(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []
  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)
  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url, count, the_word))
  return count

def googleNewsItaly():
  url = 'https://news.google.com/?hl=it&gl=IT&ceid=IT:it'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=True)
  soup = BeautifulSoup(r.text, 'html.parser')
  words = soup.find_all(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []
  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)
  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url, count, the_word))
  return count

def googleNewsChina():
    url = 'https://news.google.com/topics/CAAqJggKIiBDQkFTRWdvSkwyMHZNR1F3TlhjekVnVmxiaTFWVXlnQVAB?hl=en-US&gl=US&ceid=US:en'
    the_word = 'coronavirus'
    query = 'corona'
    r = requests.get(url, allow_redirects=True)
    soup = BeautifulSoup(r.text, 'html.parser')
    words = soup.find_all(text=lambda text: text and the_word in text)
    #print(words)
    word_list = []
    for word in words:
      wString = word.split()
      word_list.extend(wString)
    #print(word_list)
    res = [i for i in word_list if query in i]
    count = len(res)

    #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url, count, the_word))
    return count


def googleNewsKorea():
    url = 'https://news.google.com/search?hl=en-US&gl=US&ceid=US:en&q=Korea'
    the_word = 'coronavirus'
    query = 'corona'
    r = requests.get(url, allow_redirects=True)
    soup = BeautifulSoup(r.text, 'html.parser')
    words = soup.find_all(text=lambda text: text and the_word in text)
    #print(words)
    word_list = []
    for word in words:
      wString = word.split()
      word_list.extend(wString)
    #print(word_list)
    res = [i for i in word_list if query in i]
    count = len(res)

    #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url, count, the_word))
    return count

def LaRepubblica():
  url = 'https://www.repubblica.it/'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=False)
  soup = BeautifulSoup(r.text, 'html.parser')
  words = soup.find_all(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []
  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)
  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {}\ncontains {} occurrences of word: {}'.format(url, count, the_word))
  return count

def ChinaNews():
  url = 'http://global.chinadaily.com.cn/china'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=False)
  soup = BeautifulSoup(r.content, 'html.parser')
  words = soup.find_all(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []
  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)

  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {} \ncontains {} occurrences of word: {}'.format(url, count, the_word))
  return count

def HongKongNews():
  url = 'https://www.scmp.com/news/hong-kong'
  the_word = 'coronavirus'
  query = 'corona'
  r = requests.get(url, allow_redirects=False)
  soup = BeautifulSoup(r.content, 'html.parser')
  words = soup.find_all(text=lambda text: text and the_word in text)
  #print(words)
  word_list = []
  for word in words:
    wString = word.split()
    word_list.extend(wString)
  #print(word_list)

  res = [i for i in word_list if query in i]
  count = len(res)

  #print('\nUrl: {} \ncontains {} occurrences of word: {}'.format(url, count, the_word))
  return count

def appendToExcel(df):
    path = r"C:\Users\lisar\Documents\Data_Science_Exec\Corona-Tracker\Corona_Scraper.xlsx"
    writer = pd.ExcelWriter(path, engine='openpyxl')
    sheet_name = 'Sheet1'
    truncate_sheet = False
    try:
        writer.book = load_workbook(path)
        startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    df.to_excel(writer, sheet_name, startrow=startrow, header=False, index=True)
    writer.save()
    writer.close()

def execute():
  time = datetime.now()
  print('on ' + str(time))
  germany = NTV()
  england = BBC()
  usa = CNN()
  gUSA = googleNewsUSA()
  gDE = googleNewsDE()
  gItaly = googleNewsItaly()
  gChina = googleNewsChina()
  gKorea = googleNewsKorea()
  italy = LaRepubblica()
  china = ChinaNews()
  hongkong = HongKongNews()

  df = pd.DataFrame({'Date - Time': time, 'googleUSA' : [gUSA], 'googleChina': [gChina], 'googleKorea': [gKorea], 'googleDeutschland' : [gDE], 'googleItaly' : [gItaly], 'Deutschland(Focus)' : [germany], 'England(BBC)' : [england], 'USA(CNN)' : [usa], 'China(ChinaNews)' : [china], 'Italy(LaRepubblica)' : [italy], 'HongKong(HK News)' : [hongkong] })
  appendToExcel(df)
  print('done')

execute()
